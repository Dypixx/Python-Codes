# Code Created By @Dypixx

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import os

folder = os.path.expanduser('~/Documents/Book Store')
if not os.path.exists(folder):
    os.makedirs(folder)

file_name = os.path.join(folder, 'book_library.xlsx')
columns = ['Book Title', 'Author', 'User Name', 'Purchase Date', 'Return Date', 'Validity (days)', 'Penalty (₹)']
data = pd.DataFrame(columns=columns)

def load_data():
    global data
    if os.path.exists(file_name):
        data = pd.read_excel(file_name)
    else:
        data = pd.DataFrame(columns=columns)

def save_data():
    global data
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        data.to_excel(writer, index=False, sheet_name='Books')
    format_excel()

def format_excel():
    workbook = load_workbook(file_name)
    sheet = workbook.active

    header_style = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal='center')

    for cell in sheet[1]:
        cell.fill = header_style
        cell.font = header_font
        cell.alignment = header_align

    for col in sheet.columns:
        max_length = max((len(str(cell.value)) if cell.value else 0 for cell in col), default=0)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    workbook.save(file_name)

def add_book():
    global data
    print("\nAdd a New Book")
    title = input("Book Title: ").strip()
    author = input("Author: ").strip()
    user = input("User Name: ").strip()
    validity = int(input("Validity (in days): ").strip())
    penalty = float(input("Penalty (₹): ").strip())

    today = datetime.now()
    return_date = today + timedelta(days=validity)

    new_entry = {
        'Book Title': title,
        'Author': author,
        'Purchased By': user,
        'Purchase Date': today.strftime("%Y-%m-%d"),
        'Return Date': return_date.strftime("%Y-%m-%d"),
        'Validity (days)': validity,
        'Penalty (₹)': penalty
    }
    data = data._append(new_entry, ignore_index=True)
    save_data()
    print("Book added successfully!")

def display_books():
    global data
    if data.empty:
        print("No books found in the library.")
        return

    print("\nLibrary Books:")
    for idx, row in data.iterrows():
        print(f"{idx + 1}. {row['Book Title']} by {row['Author']}")

    try:
        choice = int(input("\nSelect a book number to view details (0 to exit): ").strip())
        if choice == 0:
            return
        if 1 <= choice <= len(data):
            selected = data.iloc[choice - 1]
            print("\nBook Details:")
            print(f"Title: {selected['Book Title']}")
            print(f"Author: {selected['Author']}")
            print(f"User: {selected['User Name']}")
            print(f"Purchase Date: {selected['Purchase Date']}")
            print(f"Return Date: {selected['Return Date']}")
            print(f"Validity: {selected['Validity (days)']} days")
            print(f"Penalty: ₹{selected['Penalty (₹)']}")
        else:
            print("Invalid selection.")
    except ValueError:
        print("Invalid input. Please enter a number.")

def main():
    load_data()
    while True:
        print("\nLibrary Menu")
        print("1. Add Book")
        print("2. Display Books")
        print("3. Exit")
        option = input("Choose an option: ").strip()

        if option == "1":
            add_book()
        elif option == "2":
            display_books()
        elif option == "3":
            print("Goodbye!")
            break
        else:
            print("Invalid choice. Please try again.")

main()
